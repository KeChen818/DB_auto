import re
import csv
import unicodedata
import io
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# ── Cleaning config ────────────────────────────────────────────────────────

MOJIBAKE = {
    "\u00e2\u20ac\u2122": "\u2019",  # right single quote
    "\u00e2\u20ac\u02dc": "\u2018",  # left single quote
    "\u00e2\u20ac\u0153": "\u201c",  # left double quote
    "\u00e2\u20ac\u009d": "\u201d",  # right double quote
    "\u00e2\u20ac\u201c": "\u2013",  # en-dash
    "\u00e2\u20ac\u201d": "\u2014",  # em-dash
    "\u00e2\u20ac\u00a6": "\u2026",  # ellipsis
    "\u00e2\u201a\u00ac": "\u20ac",  # euro sign
    "\u00c2\u00a3":       "\u00a3",  # pound sign
    "\u00c2\u00a9":       "\u00a9",  # copyright
    "\u00c2\u00ae":       "\u00ae",  # registered
    "\u00c2\u00b0":       "\u00b0",  # degree sign
    "\ufffd":             "",
}

HTML_ENTITIES = {
    # Standard (with semicolon)
    "&amp;": "&",  "&lt;": "<",   "&gt;": ">",
    "&quot;": '"', "&apos;": "'", "&nbsp;": " ",
    "&ndash;": "\u2013", "&mdash;": "\u2014",
    "&copy;": "\u00a9",  "&reg;": "\u00ae", "&euro;": "\u20ac",
    "&pound;": "\u00a3", "&trade;": "\u2122",
    # Malformed (without semicolon)
    "&amp": "&",  "&lt": "<",  "&gt": ">",
    "&quot": '"', "&nbsp": " ", "&ndash": "\u2013",
    "&mdash": "\u2014", "&copy": "\u00a9", "&reg": "\u00ae",
    "&euro": "\u20ac",  "&pound": "\u00a3",
}

CTRL_RE = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f\x80-\x9f"
    r"\ufeff\u200b\u200c\u200d\u2028\u2029]"
)
UNICODE_SPACES_RE = re.compile(
    r"[\u00a0\u1680\u2000-\u200a\u202f\u205f\u3000]"
)


# ── Cleaning helpers ───────────────────────────────────────────────────────

def _try_mojibake_repair(text: str) -> str:
    try:
        b = bytes(ord(c) for c in text)
        return b.decode("utf-8")
    except (ValueError, UnicodeDecodeError):
        return text


def clean_value(value) -> str:
    if not isinstance(value, str):
        return value
    value = CTRL_RE.sub("", value)
    for bad, good in MOJIBAKE.items():
        value = value.replace(bad, good)
    repaired = _try_mojibake_repair(value)
    if repaired != value:
        value = repaired
    for entity, char in HTML_ENTITIES.items():
        value = value.replace(entity, char)
    value = re.sub(r"&#x([0-9A-Fa-f]+);", lambda m: chr(int(m.group(1), 16)), value)
    value = re.sub(r"&#([0-9]+);",        lambda m: chr(int(m.group(1))),      value)
    value = UNICODE_SPACES_RE.sub(" ", value)
    value = re.sub(r" {2,}", " ", value).strip()
    value = re.sub(r"\.{4,}", "...", value)
    value = re.sub(r"\?{2,}", "?",   value)
    value = re.sub(r"!{2,}",  "!",   value)
    return unicodedata.normalize("NFC", value)


def _get_issue_type(before: str) -> str:
    if CTRL_RE.search(before):
        return "Control char"
    for key in MOJIBAKE:
        if key in before:
            return "Encoding"
    try:
        if _try_mojibake_repair(before) != before:
            return "Encoding"
    except Exception:
        pass
    for entity in HTML_ENTITIES:
        if entity in before:
            return "HTML entity"
    if re.search(r"&#[x\d]", before, re.I):
        return "HTML entity"
    if UNICODE_SPACES_RE.search(before):
        return "Non-breaking space"
    if re.search(r" {2,}", before) or before != before.strip():
        return "Extra spaces"
    return "Extra spaces"


# ── CSV helpers ────────────────────────────────────────────────────────────

def _decode_csv(file_bytes: bytes) -> tuple[str, str]:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return file_bytes.decode(enc), enc
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("latin-1", errors="replace"), "latin-1"


def _csv_rows(text: str):
    try:
        dialect = csv.Sniffer().sniff(text[:4096])
    except csv.Error:
        dialect = csv.excel
    return list(csv.reader(io.StringIO(text), dialect)), dialect


def scan_csv(file_bytes: bytes) -> tuple[list[dict], dict]:
    text, _ = _decode_csv(file_bytes)
    rows, _ = _csv_rows(text)
    headers  = rows[0] if rows else []
    issues, total = [], 0

    for r_idx, row in enumerate(rows):
        for c_idx, val in enumerate(row):
            if not val:
                continue
            total += 1
            cleaned = clean_value(val)
            if cleaned != val:
                col_label = headers[c_idx] if r_idx > 0 and c_idx < len(headers) else f"Col {c_idx + 1}"
                issues.append({
                    "Sheet":  "CSV",
                    "Cell":   f"R{r_idx + 1} / {col_label}",
                    "Issue":  _get_issue_type(val),
                    "Before": val,
                    "After":  cleaned,
                    "_r": r_idx,
                    "_c": c_idx,
                })

    return issues, {"cells_scanned": total, "issues_found": len(issues), "sheets": 1}


def apply_fixes_csv(file_bytes: bytes, issues: list[dict]) -> bytes:
    text, _  = _decode_csv(file_bytes)
    rows, dialect = _csv_rows(text)
    fix_map  = {(r["_r"], r["_c"]): r["After"] for r in issues}

    for (r_idx, c_idx), after in fix_map.items():
        if r_idx < len(rows) and c_idx < len(rows[r_idx]):
            rows[r_idx][c_idx] = after

    buf = io.StringIO()
    csv.writer(buf, dialect=dialect).writerows(rows)
    return buf.getvalue().encode("utf-8-sig")   # BOM makes Excel open UTF-8 CSV correctly


# ── Excel helpers ──────────────────────────────────────────────────────────

def scan_excel(file_bytes: bytes) -> tuple[list[dict], dict]:
    wb = load_workbook(io.BytesIO(file_bytes))
    issues, total = [], 0

    for sheet_name in wb.sheetnames:
        for row in wb[sheet_name].iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                total += 1
                cleaned = clean_value(cell.value)
                if cleaned != cell.value:
                    issues.append({
                        "Sheet":  sheet_name,
                        "Cell":   cell.coordinate,
                        "Issue":  _get_issue_type(cell.value),
                        "Before": cell.value,
                        "After":  cleaned,
                    })

    return issues, {"cells_scanned": total, "issues_found": len(issues), "sheets": len(wb.sheetnames)}


def apply_fixes_excel(file_bytes: bytes, issues: list[dict]) -> bytes:
    wb      = load_workbook(io.BytesIO(file_bytes))
    fix_map = {(r["Sheet"], r["Cell"]): r["After"] for r in issues}

    for sheet_name in wb.sheetnames:
        for row in wb[sheet_name].iter_rows():
            for cell in row:
                key = (sheet_name, cell.coordinate)
                if key in fix_map:
                    cell.value = fix_map[key]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Streamlit UI ───────────────────────────────────────────────────────────

st.set_page_config(page_title="Excel / CSV Cleaner", page_icon="\U0001f9f9", layout="wide")
st.title("\U0001f9f9 Excel / CSV Cleaner")
st.caption("Scan, review, and fix encoding errors and whitespace issues in Excel or CSV files.")

# ── Step 1: Upload ─────────────────────────────────────────────────────────
uploaded = st.file_uploader(
    "Upload a file",
    type=["xlsx", "xls", "csv"],
    label_visibility="collapsed",
)

if not uploaded:
    st.info("Drop an .xlsx, .xls, or .csv file above to get started.")
    st.stop()

file_bytes = uploaded.read()
is_csv     = uploaded.name.lower().endswith(".csv")
file_type  = "CSV" if is_csv else "Excel"
st.success(f"**{uploaded.name}** ({file_type}) \u2014 {len(file_bytes) / 1024:.1f} KB loaded")

# ── Step 2: Run Check ──────────────────────────────────────────────────────
if st.button("\u25b6 Run Check", type="primary"):
    with st.spinner("Scanning cells\u2026"):
        if is_csv:
            issues, stats = scan_csv(file_bytes)
        else:
            issues, stats = scan_excel(file_bytes)
    st.session_state.update({
        "issues":     issues,
        "stats":      stats,
        "file_bytes": file_bytes,
        "file_name":  uploaded.name,
        "is_csv":     is_csv,
    })

# ── Step 3: Preview ────────────────────────────────────────────────────────
if "issues" not in st.session_state:
    st.stop()

issues     = st.session_state["issues"]
stats      = st.session_state["stats"]
file_bytes = st.session_state["file_bytes"]
file_name  = st.session_state["file_name"]
is_csv     = st.session_state["is_csv"]

col1, col2, col3 = st.columns(3)
col1.metric("Cells scanned",  f"{stats['cells_scanned']:,}")
col2.metric("Issues found",   f"{stats['issues_found']:,}")
col3.metric("Sheets / files", stats["sheets"])

if not issues:
    st.success("\u2705 No issues found \u2014 your file is already clean!")

# ── Debug panel ────────────────────────────────────────────────────────────
with st.expander("\U0001f50d Debug: inspect raw cell values"):
    st.caption("Shows up to 30 non-empty string cells exactly as read. "
               "repr() reveals invisible characters (\\xa0, \\t, etc.).")
    samples = []

    if is_csv:
        text, enc = _decode_csv(file_bytes)
        st.caption(f"Detected encoding: **{enc}**")
        rows, _ = _csv_rows(text)
        headers = rows[0] if rows else []
        for r_idx, row in enumerate(rows):
            for c_idx, val in enumerate(row):
                if not val:
                    continue
                cleaned = clean_value(val)
                col_label = headers[c_idx] if r_idx > 0 and c_idx < len(headers) else f"Col {c_idx+1}"
                samples.append({
                    "Location":    f"R{r_idx+1} / {col_label}",
                    "repr(value)": repr(val),
                    "Would fix?":  "\u2705 yes" if cleaned != val else "\u2014 no",
                    "Proposed":    repr(cleaned) if cleaned != val else "",
                })
                if len(samples) >= 30:
                    break
            if len(samples) >= 30:
                break
    else:
        wb_dbg = load_workbook(io.BytesIO(file_bytes))
        for sn in wb_dbg.sheetnames:
            for row in wb_dbg[sn].iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str) or not cell.value:
                        continue
                    cleaned = clean_value(cell.value)
                    samples.append({
                        "Location":    f"{sn} / {cell.coordinate}",
                        "repr(value)": repr(cell.value),
                        "Would fix?":  "\u2705 yes" if cleaned != cell.value else "\u2014 no",
                        "Proposed":    repr(cleaned) if cleaned != cell.value else "",
                    })
                    if len(samples) >= 30:
                        break
                if len(samples) >= 30:
                    break
            if len(samples) >= 30:
                break

    if samples:
        st.dataframe(pd.DataFrame(samples), use_container_width=True, hide_index=True)
    else:
        st.warning("No string cells found in this file.")

# ── Manual test ────────────────────────────────────────────────────────────
st.divider()
st.subheader("Test a value manually")
st.caption("Paste any cell content here to see exactly what the cleaner does to it.")
test_val = st.text_input("Cell value to test", placeholder="e.g.  hello   world  ")
if test_val:
    result = clean_value(test_val)
    c1, c2 = st.columns(2)
    c1.text_area("Input repr()",  repr(test_val), height=80)
    c2.text_area("Output repr()", repr(result),   height=80)
    if result != test_val:
        st.success("\u2705 Cleaner would change this cell.")
    else:
        st.warning("\u26a0 No change \u2014 already clean or character type not yet supported.")

if not issues:
    st.stop()

st.divider()

# ── Issue table ────────────────────────────────────────────────────────────
all_types = sorted({r["Issue"] for r in issues})
tabs = st.tabs(["All"] + all_types)

def show_table(rows):
    if not rows:
        st.info("No issues of this type.")
        return
    df = pd.DataFrame(rows)[["Sheet", "Cell", "Issue", "Before", "After"]]
    st.dataframe(
        df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Before": st.column_config.TextColumn("Before",           width="large"),
            "After":  st.column_config.TextColumn("After (proposed)", width="large"),
        },
    )

with tabs[0]:
    show_table(issues)
for i, t in enumerate(all_types, 1):
    with tabs[i]:
        show_table([r for r in issues if r["Issue"] == t])

st.divider()

# ── Step 4: Accept / Discard ───────────────────────────────────────────────
st.subheader("Apply changes?")
st.caption(f"This will fix all **{len(issues)}** identified issues and prepare a cleaned file for download.")

accept_col, discard_col, _ = st.columns([1, 1, 4])

with accept_col:
    if st.button("\u2705 Accept & download", type="primary", use_container_width=True):
        with st.spinner("Applying fixes\u2026"):
            if is_csv:
                clean_bytes = apply_fixes_csv(file_bytes, issues)
                clean_name  = re.sub(r"\.csv$", "_cleaned.csv", file_name, flags=re.I)
                mime        = "text/csv"
            else:
                clean_bytes = apply_fixes_excel(file_bytes, issues)
                clean_name  = re.sub(r"\.(xlsx?)$", r"_cleaned.\1", file_name, flags=re.I)
                mime        = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        st.download_button(
            label="\u2b07 Download clean file",
            data=clean_bytes,
            file_name=clean_name,
            mime=mime,
            use_container_width=True,
        )

with discard_col:
    if st.button("\u2716 Discard", use_container_width=True):
        for key in ["issues", "stats", "file_bytes", "file_name", "is_csv"]:
            st.session_state.pop(key, None)
        st.rerun()
